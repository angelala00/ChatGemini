import os
import textract
import imghdr
from openpyxl import load_workbook
from docx import Document
from PyPDF2 import PdfReader
from app.utils.model_tool import convert_image_message, MODEL_NAME_VL


async def extract_text_from_file(file_path: str, file_type: str):
    if not os.path.exists(file_path):
        raise FileNotFoundError("FilePath not found")

    # if file_type == '.xlsx':
    #     result = parse_excel(file_path)
    #     return {"text": result}
    # elif file_type == '.docx' or file_type == '.doc':
    #     result = parse_word(file_path)
    #     return {"text": result}
    # elif file_type == '.pdf':
    #     result = parse_pdf(file_path)
    #     return {"text": result}
    # else:
    #     raise Exception(f"UnSupport file type:{file_type}")

    if imghdr.what(file_path) is not None:
        from app.chat_service import _ask_once_stream

        query = "提取图片信息"
        messages = [{"role": "user", "content": convert_image_message(file_path, query)}]
        # print(f"messages:{messages}")
        text_content = ""
        async for event in _ask_once_stream(messages, None, MODEL_NAME_VL):
            if event.get("type") == "text.delta":
                text_content += event.get("data")["text"]
        return text_content
    else:
        raw = textract.process(file_path, extension=file_type)
        text = raw.decode('utf-8')
    return text


def parse_excel(file_path):
    """解析Excel文件"""
    wb = load_workbook(file_path)
    sheet = wb.active
    text = ""
    for row in sheet.iter_rows(values_only=True):
        text += " ".join(str(cell) for cell in row) + "\n"
    return text


def parse_word(file_path):
    """解析Word文件"""
    doc = Document(file_path)
    text = ""
    for paragraph in doc.paragraphs:
        text += paragraph.text + "\n"
    return text


def parse_pdf(file_path):
    """解析Pdf文件"""
    reader = PdfReader(file_path)
    text = ""
    for page in reader.pages:
        # 提取页面中的文字
        text += page.extract_text()
    return text
